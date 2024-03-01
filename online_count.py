from openpyxl import load_workbook
import pathlib
import sys

print("街乡平台设备数据情况比对程序,数据表文件名必须以'街乡平台设备数据情况比对表'开头，以'.xlsx'结尾,且在线原数据位于第三个sheet;\n统计结果文件名必须为'小区统计表1775.xlsx'")
file_list = [file for file in pathlib.Path('.').iterdir() if file.name.startswith("街乡平台设备数据情况比对表") and file.name.endswith(".xlsx")]

for file_name in file_list:
    workbook = load_workbook(file_name,read_only=True,data_only=True)
    sheet = workbook.worksheets[2]  # 获取第三个sheet
    result = {}
    # 遍历B列和D列并统计在线总数
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4, values_only=True):
        category, status = str(row[0]).split('/')[-1], row[2]    # 所选区域-->小区名称，设备状态
        if category not in result:
            result[category] = "全部离线"
        if status == '在线':
            if result.get(category, 0) == "全部离线":
                result[category] = 0
            result[category] = result.get(category, 0) + 1
# 从"小区统计表1775.xlsx"
file = load_workbook("小区统计表1775.xlsx",data_only=True,read_only=False)
sheet = file.active
# 遍历E列内容，从第二行开始
for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value in result:
            # 如果找到，则设置为字典value的值
            sheet.cell(row=cell.row, column=11, value=result[cell.value])
            print(cell.value, result[cell.value])
        else:
            # 如果找不到，将对应K列单元格值设置0
            sheet.cell(row=cell.row, column=11, value="平台无设备")
try:
    file.save("小区统计表1775.xlsx")
except PermissionError:
    print("结果写入失败,请关闭文件后重试")
    sys.exit(1)

input("结果写入成功,按enter退出...")
sys.exit(0)