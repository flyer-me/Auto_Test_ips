# -*- coding: utf-8 -*-
import contextlib
import os
import sys
from datetime import datetime
from ipaddress import ip_address

from chardet import detect
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import configparser



'''
description: 
param {*} filename
param {*} header_row
return {*}
'''
def is_sheet_empty(sheet):
    return sheet.max_row == 1 and sheet.max_column == 1


'''
description: 
param {*} filename
param {*} sheet
return {*}
'''
def get_ip_row_in_sheet(filename, sheet):
    if is_sheet_empty(sheet):
        return -1
    global backnames
    row_generator = sheet.iter_rows(values_only=True)
    for _ in range(5, 0, -1):
        try:
            header_row = next(row_generator)
        except StopIteration:
            return -1
        for back_name in backnames:
            with contextlib.suppress(ValueError):
                return header_row.index(back_name)
    return -1


'''
description: 
param {*} files
return {*}
'''
def read_xlsx(files=None):
    data = []
    for fn in files:
        try:
            wb = load_workbook(fn, read_only=True)
        except PermissionError:
            input("文件被占用,无法读取,关闭后按enter继续")

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(os.path.basename(fn), sheet)
            if col_num <= -1:
                continue
            for row in sheet.iter_rows():
                cell_value = row[col_num].value
                if cell_value is not None:
                    with contextlib.suppress(ValueError):
                        ip_address(cell_value)
                        data.append(cell_value)
    return list(set(data))


def get_ip_list(xfile, result):
    data = read_xlsx(xfile)
    with open(result, 'w') as f:
        for item in data:
            f.write(f"{item}\n")


def count_lines(filename):
    with open(filename, 'r') as file:
        count = sum(1 for line in file if line.strip())
    return count

'''
description: 
param {*} tool_path
param {*} xfile
param {*} result
return {*}
'''
def call_ping(tool_path, xfile, result, timeout = 5000, size = 4):
    if not os.path.exists(result):
        with open(result, 'w') as f:
            f.write('')
    path = os.path.abspath(tool_path)
    command = f'{path} /loadfile {xfile} /stab {result} /PingTimeout {timeout} /PingSize {size}'
    print(f'超时IP数{count_lines(xfile)}.')
    os.system(command)


def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    if os.path.exists(file):
        with open(file, 'rb') as f:
            data = f.read()
            return detect(data)['encoding']


'''
description: 
param {*} ip_list
param {*} bad_ip_list
return {*}
'''
def get_bad_ip(ip_list, bad_ip_list):
    with open(ip_list, 'r', encoding=get_encoding(ip_list)) as f:
        lines = f.readlines()
    with open(bad_ip_list, 'w') as f:
        for line in lines:
            parts = line.split()
            if parts[1] == '0':
                f.write(parts[0] + '\n')


'''
description: 
param {*} file_list
param {*} bad_ip
return {*}
    规则: 在线统计结果放置于表格右侧空白一列
'''
def write_result(file_list, ip_file, is_in, not_in):
    for fn in file_list:
        try:
            wb = load_workbook(fn)
        except PermissionError:
            input("文件被占用,无法读取,关闭后按enter继续")

        for sheet in wb.worksheets:
            col_num = get_ip_row_in_sheet(os.path.basename(fn), sheet)
            if col_num <= -1:
                continue
            max_col = sheet.max_column
            new_col = max_col + 1  # 增加在线统计结果列
            sheet.cell(row=1, column=new_col).value = datetime.now().strftime("%m%d %H:%M")

            with open(ip_file, "r") as f:
                ip_list = set(f.read().splitlines())

            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_num + 1).value  # sheet.cell 的 column 是从1开始,而col_num是从0开始
                try:
                    ip_address(cell_value)
                except ValueError:
                    continue
                #if ip_address(cell_value).is_private == False:
                if cell_value in ip_list:
                    sheet.cell(row=row, column=new_col).value = is_in
                else:
                    sheet.cell(row=row, column=new_col).value = not_in
                #else:
                #    sheet.cell(row=row, column=new_col).value = ''
        wb.save(fn)

'''
description: 
param {*} write_index_offset 写入结果列与IP地址所在列的index差值
    定制规则：
        if cell_value in ip_list 的结果记为I,IP单元格的右3 cell记为R,左3 cell记为L
        1.R根据I正常填充is_in、not_in
        2.L填充红色,跳过
        3.L无填充色/白色且非空,I为真,L.value=is_in
        4.L无填充色/白色,且空;I为假,L.value=1,填充红色
'''
def write_result_color(file_list, ip_file, is_in, not_in, write_index_offset=3):
    fille = PatternFill("solid",fgColor="FFFF0000")
    for fn in file_list:
        try:
            wb = load_workbook(fn)
        except PermissionError:
            input("文件被占用,无法读取,关闭后按enter继续")

        for sheet in wb.worksheets:     # 跳过无IP的sheet
            col_num = get_ip_row_in_sheet(os.path.basename(fn), sheet)
            if col_num <= -1:
                continue
            
            with open(ip_file, "r") as f:
                ip_list = set(f.read().splitlines())

            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_num + 1).value  # sheet.cell 的 column 是从1开始,而col_num是从0开始
                try:
                    ip_address(cell_value)
                except ValueError:
                    continue
                
                bool_is_in = cell_value in ip_list
                sheet.cell(row=row, column=col_num + 1 + write_index_offset).value  = is_in if bool_is_in else not_in

                fill = sheet.cell(row=row, column=col_num - 2).fill     #涂色位置
                value = sheet.cell(row=row, column=col_num - 2).value   #涂色格的值

                if fill.fgColor.rgb != 'FFFF0000':          #标准红色跳过
                    if value is not None:                     #非红,非空
                        sheet.cell(row=row, column=col_num - 2).value = is_in if bool_is_in else 1
                    elif not bool_is_in:      #非红,不在list中,空
                        sheet.cell(row=row, column=col_num - 2).value = 1
                        sheet.cell(row=row, column=col_num - 2).fill = fille
                        print(cell_value," 新增在线,加入未上墙")
        wb.save(fn)

def remove_txt():
    file_path = os.getcwd()
    file_ext = '.txt'
    for path in os.listdir(file_path):
        path_list = os.path.join(file_path, path)
        if os.path.isfile(path_list) and (os.path.splitext(path_list)[1]) == file_ext:
            os.remove(path_list)

'''
description: 对XLSX文件列表,执行IP连通性测试,并回写结果
param {*} file_list
return {*}
'''
def ip_xlsx_test(file_list, ping_timeout, ping_size, times):
    get_ip_list(file_list, "ip.txt")
    bad_name = "bad.txt"
    with open("ip.txt", 'r') as file1, open(bad_name, 'w') as file2:
        file2.write(file1.read())

    print(f'ping设置:超时{ping_timeout}ms,{ping_size}字节')
    while(times > 0):
        call_ping(r'tools\PingInfoView.exe', bad_name, "result.txt", ping_timeout, ping_size)
        get_bad_ip("result.txt", bad_name)
        times = times - 1
        print(f'剩余{times}次,',end='')

    # 结果回写
    # write_result(file_list, bad_name, "超时", "成功")
    write_result_color(file_list, bad_name, "掉线", "正常")

def main():
    global config
    ping_timeout = config.getint('cfg', 'ping_timeout', fallback=5000)    #ping超时时间
    ping_times = config.getint('cfg', 'ping_times', fallback=25)        #ping尝试次数
    file_list = [f for f in os.listdir() if f.endswith('.xlsx')]
    ip_xlsx_test(file_list, ping_timeout, 4, ping_times)
    remove_txt()

if __name__ == '__main__':
    print(os.getcwd())
    config = configparser.ConfigParser()
    config.read('config/config.ini',encoding="utf-8-sig")
    head_name = config.get('cfg', 'head_name', fallback="IP")        #定义表头
    backnames = ['专网IP','IP地址']
    backnames.insert(0,head_name)
    try:
        main()
        input("操作完成,按enter退出")
        sys.exit(0)
    except Exception as e:
        print(e)
